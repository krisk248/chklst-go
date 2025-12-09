"""
Excel export service using pandas and openpyxl
"""
import pandas as pd
from datetime import datetime
from typing import List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import io


class ExcelService:
    """Service for generating Excel exports"""

    def __init__(self):
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    def generate_deployment_report(
        self,
        deployments: List[Dict[str, Any]],
        month: int,
        year: int
    ) -> bytes:
        """
        Generate Excel report for deployments

        Args:
            deployments: List of deployment dictionaries
            month: Report month (1-12)
            year: Report year

        Returns:
            Excel file as bytes
        """
        # Convert to DataFrame
        df = pd.DataFrame(deployments)

        if df.empty:
            # Create empty report
            df = pd.DataFrame(columns=[
                'ID', 'Jira ID', 'Project', 'Component', 'Developer',
                'Build Server', 'Deploy Server', 'Environment',
                'Timestamp', 'Status'
            ])

        # Format timestamp
        if 'timestamp' in df.columns and not df.empty:
            df['timestamp'] = pd.to_datetime(df['timestamp']).dt.strftime('%Y-%m-%d %H:%M:%S')

        # Reorder and rename columns
        column_mapping = {
            'id': 'ID',
            'jira_id': 'Jira ID',
            'project_name': 'Project',
            'component_name': 'Component',
            'developer': 'Developer',
            'build_server': 'Build Server',
            'deploy_server': 'Deploy Server',
            'environment': 'Environment',
            'timestamp': 'Timestamp',
            'status': 'Status',
            'notes': 'Notes'
        }

        # Select and rename available columns
        available_cols = [col for col in column_mapping.keys() if col in df.columns]
        df_export = df[available_cols].copy()
        df_export.rename(columns=column_mapping, inplace=True)

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = f"Deployments {year}-{month:02d}"

        # Add title row
        title = f"Deployment Report - {datetime(year, month, 1).strftime('%B %Y')}"
        ws.merge_cells('A1:F1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')

        # Add metadata
        ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(italic=True, size=10)

        # Add data starting from row 4
        start_row = 4

        # Write headers
        for col_num, column_name in enumerate(df_export.columns, 1):
            cell = ws.cell(row=start_row, column=col_num)
            cell.value = column_name
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = self.border

        # Write data
        for row_num, row_data in enumerate(df_export.values, start_row + 1):
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = self.border
                cell.alignment = Alignment(vertical='center')

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze header row
        ws.freeze_panes = 'A5'

        # Add summary sheet
        summary_ws = wb.create_sheet("Summary")

        # Summary statistics
        summary_data = [
            ["Metric", "Value"],
            ["Total Deployments", len(deployments)],
            ["Report Period", f"{datetime(year, month, 1).strftime('%B %Y')}"],
            ["Generated", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ]

        # Add deployment counts by environment if data exists
        if 'environment' in df.columns and not df.empty:
            env_counts = df['environment'].value_counts()
            summary_data.append(["", ""])
            summary_data.append(["Deployments by Environment", ""])
            for env, count in env_counts.items():
                summary_data.append([env, count])

        # Write summary data
        for row_num, (label, value) in enumerate(summary_data, 1):
            summary_ws.cell(row=row_num, column=1, value=label).font = Font(bold=True)
            summary_ws.cell(row=row_num, column=2, value=value)

        summary_ws.column_dimensions['A'].width = 30
        summary_ws.column_dimensions['B'].width = 20

        # Save to bytes
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer.getvalue()

    def generate_statistics_report(
        self,
        stats: Dict[str, Any],
        month: int,
        year: int
    ) -> bytes:
        """
        Generate Excel report for deployment statistics

        Args:
            stats: Statistics dictionary
            month: Report month
            year: Report year

        Returns:
            Excel file as bytes
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Statistics"

        # Title
        ws['A1'] = f"Deployment Statistics - {datetime(year, month, 1).strftime('%B %Y')}"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')

        # Statistics data
        row = 3
        for key, value in stats.items():
            ws.cell(row=row, column=1, value=key.replace('_', ' ').title()).font = Font(bold=True)
            ws.cell(row=row, column=2, value=value)
            row += 1

        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20

        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)

        return excel_buffer.getvalue()
